# Trade journal → spreadsheet.
#
# Schema (36 cols):
#   #, Date, Symbol, Dir, Entry, SL, SL%, T1, T1%, T2, T2%, R:R, ATR, Conf%,
#   Size₹, Risk₹, Tech, News, Sector, Macro, Earn, Fund, Opts,
#   Why Bought(AI), ETA,
#   Live ₹ (*), Live P&L ₹ (*), Live %, Days Held,   ← Google Sheets formulas
#   Status, Target Hit, Closed, Duration, P&L ₹, P&L %, Expert Note(AI)
#
# (*) Live ₹ uses =GOOGLEFINANCE("NSE:SYMBOL","price") — auto-refreshes ~5 min.
#     Live P&L auto-recalculates from live price while Status = OPEN.
#
# Idempotent sync: new open-rows append; closed-rows update their close-columns.

from __future__ import annotations

import asyncio
import os
import threading
from datetime import datetime, timedelta, timezone

from sqlalchemy import select, or_
from sqlalchemy.ext.asyncio import AsyncSession

from db.models import PaperTrade, MasterIntelligenceScore, TradeStatus, AgentTrade, AgentDecision, OpenPosition
from integrations.trade_explainer import (
    build_expert_note, build_hold_analysis, build_postmortem_note, estimate_eta_to_target,
)
from utils.config import settings
from utils.logger import logger

_IST       = timezone(timedelta(hours=5, minutes=30))
_FILE_LOCK = threading.Lock()

# ── Column schema ─────────────────────────────────────────────────────────────
SCHEMA: list[tuple[str, str]] = [
    # ── Trade identity ─────────────────────────────────────────────────────
    ("trade_id",      "#"),
    ("opened_at",     "Bought (IST)"),
    ("symbol",        "Symbol"),
    # ── F&O detail (EQUITY for cash; populated for options/futures) ─────────
    ("instrument",    "Instrument"),
    ("strike",        "Strike"),
    ("opt_type",      "CE/PE"),
    ("expiry",        "Expiry"),
    ("direction",     "Dir"),
    ("qty",           "Qty"),
    # ── Levels ─────────────────────────────────────────────────────────────
    ("entry",         "Entry ₹"),
    ("stop",          "SL ₹"),
    ("stop_pct",      "SL%"),
    ("target_1",      "T1 ₹"),
    ("t1_pct",        "T1%"),
    ("target_2",      "T2 ₹"),
    ("t2_pct",        "T2%"),
    ("rr",            "R:R"),
    ("atr",           "ATR"),
    # ── Signal quality ─────────────────────────────────────────────────────
    ("confidence",    "Conf %"),
    ("size_rs",       "Size ₹"),
    ("max_risk_rs",   "Max Risk ₹"),
    # ── Hub 7-factor breakdown ──────────────────────────────────────────────
    ("hub_tech",      "Tech"),
    ("hub_news",      "News"),
    ("hub_sector",    "Sector"),
    ("hub_macro",     "Macro"),
    ("hub_earn",      "Earn"),
    ("hub_fund",      "Fund"),
    ("hub_opts",      "Opts"),
    # ── AI context ─────────────────────────────────────────────────────────
    ("why_bought",    "Why Bought (AI)"),
    ("eta_t1",        "ETA to T1"),
    # ── LIVE data (Google Sheets formulas — auto-refresh) ──────────────────
    ("live_price",    "Live ₹ 🔴"),
    ("live_pnl_rs",   "Live P&L ₹"),
    ("live_pnl_pct",  "Live P&L %"),
    ("days_held",     "Days Held"),
    # ── Outcome (filled when trade closes) ─────────────────────────────────
    ("status",        "Status"),
    ("target_hit",    "Target Hit"),
    ("closed_at",     "Closed (IST)"),
    ("duration",      "Duration"),
    ("pnl_rs",        "Final P&L ₹"),
    ("pnl_pct",       "Final P&L %"),
    ("expert_note",   "Expert Note (AI)"),
]
HEADERS   = [h for _, h in SCHEMA]
KEYS      = [k for k, _ in SCHEMA]
CLOSE_KEYS = {"status", "target_hit", "closed_at", "duration", "pnl_rs", "pnl_pct", "expert_note"}

_CI  = {k: i for i, k in enumerate(KEYS)}   # key → 0-based column index
_COL = {}                                    # key → column letter (filled after _col_letter defined)

def _col_letter(idx: int) -> str:
    s, i = "", idx + 1
    while i:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s

# Pre-compute column letters for formula use
for _k, _i in _CI.items():
    _COL[_k] = _col_letter(_i)

# Column widths (pixels) – 36 cols
_COL_WIDTHS = [
    50, 130, 110, 50,           # #, Date, Symbol, Dir
    80, 80, 55, 80, 55, 80, 55, 50, 60,   # Entry→ATR
    60, 90, 90,                  # Conf, Size, Risk
    50, 50, 55, 55, 50, 50, 50,  # Hub 7
    340, 130,                    # Why, ETA
    85, 100, 80, 75,             # Live: price, P&L₹, P&L%, Days
    80, 170, 130, 80, 90, 75, 340,  # Status→Expert
]

# ── Colour palette ─────────────────────────────────────────────────────────────
def _rgb(r, g, b):
    return {"red": r/255, "green": g/255, "blue": b/255}

C_HEADER_BG   = _rgb(15,  23,  42)   # slate-900
C_HEADER_FG   = _rgb(248, 250, 252)
C_BUY_BG      = _rgb(240, 253, 244)  # green-50
C_SELL_BG     = _rgb(255, 241, 242)  # rose-50
C_LIVE_HEADER = _rgb(6,   182, 212)  # cyan accent for live section
C_HUB_POS     = _rgb(220, 252, 231)
C_HUB_NEG     = _rgb(254, 226, 226)
C_STATUS_OPEN = _rgb(254, 243, 199)
C_STATUS_T1   = _rgb(209, 250, 229)
C_STATUS_STOP = _rgb(254, 226, 226)
C_PNL_POS     = _rgb(209, 250, 229)
C_PNL_NEG     = _rgb(254, 226, 226)
C_SECTION_BG  = _rgb(248, 250, 252)

# ── Helpers ───────────────────────────────────────────────────────────────────
def _ist(dt):
    if not dt:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(_IST).strftime("%Y-%m-%d %H:%M")

def _fmt_duration(start, end):
    secs = max(0, (end - start).total_seconds())
    days, rem = divmod(int(secs), 86400)
    hours = rem // 3600
    if days:
        return f"{days}d {hours}h"
    mins = (rem % 3600) // 60
    return f"{hours}h {mins}m" if hours else f"{mins}m"

def _hub_scores(mis):
    if not mis:
        return {}
    return {
        "hub_tech":   round(mis.technical_score),
        "hub_news":   round(mis.news_score),
        "hub_sector": round(mis.sector_score),
        "hub_macro":  round(mis.macro_score),
        "hub_earn":   round(mis.earnings_score),
        "hub_fund":   round(mis.fundamental_score),
        "hub_opts":   round(mis.options_score),
    }

def _trade_mgmt(trade):
    snap = trade.indicator_snapshot or {}
    return snap.get("trade_mgmt", {}) if isinstance(snap, dict) else {}

def _pct(a, b):
    try:
        return round((a - b) / b * 100, 2) if b else 0.0
    except Exception:
        return 0.0

def _rr(entry, stop, t1):
    try:
        risk = abs(entry - stop)
        return round(abs(t1 - entry) / risk, 2) if risk else 0.0
    except Exception:
        return 0.0

# ── Google Sheets live formulas ────────────────────────────────────────────────
# These are written as formula strings; gspread interprets them when
# value_input_option="USER_ENTERED". {R} is replaced with the actual row number.

def _live_formulas(row_num="{ROW}") -> dict:
    """Return formula strings for the 4 live-data columns.

    row_num may be a concrete int or the default sentinel "{ROW}".
    GoogleSheetsSink.append() replaces "{ROW}" with the actual row number.
    """
    R = row_num
    C  = _COL                      # shorthand
    # Symbol in our sheet has .NS stripped → "DALBHARAT", "RELIANCE" etc.
    live_col  = C["live_price"]    # Z
    entry_col = C["entry"]         # E
    dir_col   = C["direction"]     # D
    size_col  = C["size_rs"]       # O
    stat_col  = C["status"]        # AD
    pnl_col   = C["pnl_rs"]       # AH
    pnl_pct_c = C["pnl_pct"]      # AI
    cl_col    = C["closed_at"]     # AF
    op_col    = C["opened_at"]     # B

    return {
        # Live price via GOOGLEFINANCE (auto-refreshes ~5 min, 15-min delay on free tier)
        "live_price": (
            f'=IFERROR(GOOGLEFINANCE("NSE:"&{C["symbol"]}{R},"price"),"⏳")'
        ),
        # Live P&L ₹ — shows unrealized while OPEN, final once closed
        "live_pnl_rs": (
            f'=IFERROR(IF(ISNUMBER({pnl_col}{R}),{pnl_col}{R},'
            f'IF({live_col}{R}="","—",'
            f'IF({dir_col}{R}="SELL",'
            f'({entry_col}{R}-{live_col}{R})*{size_col}{R}/{entry_col}{R},'
            f'({live_col}{R}-{entry_col}{R})*{size_col}{R}/{entry_col}{R}'
            f'))),"—")'
        ),
        # Live P&L % (same logic, returns percentage)
        "live_pnl_pct": (
            f'=IFERROR(IF(ISNUMBER({pnl_col}{R}),{pnl_pct_c}{R},'
            f'IF({live_col}{R}="","—",'
            f'IF({dir_col}{R}="SELL",'
            f'({entry_col}{R}-{live_col}{R})/{entry_col}{R}*100,'
            f'({live_col}{R}-{entry_col}{R})/{entry_col}{R}*100'
            f'))),"—")'
        ),
        # Days held (integer)
        "days_held": (
            f'=IFERROR(IF({cl_col}{R}<>"",'
            f'DAYS(LEFT({cl_col}{R},10),LEFT({op_col}{R},10)),'
            f'DAYS(TEXT(TODAY(),"yyyy-mm-dd"),LEFT({op_col}{R},10))),"—")'
        ),
    }

# ── Row builders ──────────────────────────────────────────────────────────────
def _open_row(trade: PaperTrade, hub) -> dict:
    mgmt      = _trade_mgmt(trade)
    entry     = trade.entry_price
    stop      = trade.stop_loss
    t1        = mgmt.get("target_1", trade.take_profit)
    t2        = mgmt.get("target_2", trade.take_profit)
    atr       = mgmt.get("atr", 0.0)
    direction = trade.direction.value
    units     = trade.size_units
    size_rs   = round(entry * units, 0)
    max_risk  = round(abs(entry - stop) * units, 0)
    hub_scores = _hub_scores(hub)
    hub_note   = {k.replace("hub_", ""): v for k, v in hub_scores.items()} if hub_scores else None
    note = build_expert_note(trade.symbol, direction, entry, stop, t1, t2,
                             trade.signal_confidence, hub_note, trade.ai_reason)
    itype = getattr(trade, "instrument_type", "EQUITY") or "EQUITY"
    is_fno = itype in ("CE", "PE", "FUTURE")
    row = {
        "trade_id":    trade.id,
        "opened_at":   _ist(trade.opened_at),
        "symbol":      (getattr(trade, "underlying_symbol", None) or trade.symbol.replace(".NS", "")) if is_fno else trade.symbol.replace(".NS", ""),
        # F&O detail
        "instrument":  "FUT" if itype == "FUTURE" else (itype if itype in ("CE", "PE") else "EQUITY"),
        "strike":      round(getattr(trade, "strike_price", 0) or 0, 0) if getattr(trade, "strike_price", None) else "",
        "opt_type":    getattr(trade, "option_type", "") or "",
        "expiry":      trade.expiry_date.strftime("%d-%b-%Y") if getattr(trade, "expiry_date", None) else "",
        "direction":   direction,
        "qty":         int(round(units)),
        "entry":       round(entry, 2),
        "stop":        round(stop, 2),
        "stop_pct":    f"{_pct(stop, entry):+.1f}%",
        "target_1":    round(t1, 2),
        "t1_pct":      f"{_pct(t1, entry):+.1f}%",
        "target_2":    round(t2, 2),
        "t2_pct":      f"{_pct(t2, entry):+.1f}%",
        "rr":          f"1:{_rr(entry, stop, t1)}",
        "atr":         round(atr, 2),
        "confidence":  round(trade.signal_confidence, 1),
        "size_rs":     size_rs,
        "max_risk_rs": max_risk,
        "why_bought":  note,
        "eta_t1":      estimate_eta_to_target(entry, t1, atr, direction),
        "status":      "OPEN",
    }
    row.update(hub_scores)
    # Always add live formula templates; GoogleSheetsSink.append() substitutes
    # {ROW} with the concrete row number. LocalExcelSink strips =... strings.
    row.update(_live_formulas())
    return row


def _close_partial(trade: PaperTrade) -> dict:
    mgmt   = _trade_mgmt(trade)
    t1     = mgmt.get("target_1", trade.take_profit)
    exit_p = trade.exit_price or 0.0
    if trade.status == TradeStatus.STOPPED:
        status, hit = "STOPPED ❌", "Stopped out"
    else:
        hit_t1 = (exit_p >= t1) if trade.direction.value == "BUY" else (exit_p <= t1)
        if hit_t1:
            status, hit = "T1 HIT ✅", f"T1 ₹{t1:,.2f} reached"
        else:
            status, hit = "CLOSED", "Closed before T1"
    dur  = _fmt_duration(trade.opened_at, trade.closed_at) if trade.closed_at else ""
    pnl  = trade.pnl or 0.0
    pct  = trade.pnl_percent or 0.0
    note = build_postmortem_note(trade.symbol, trade.direction.value, trade.entry_price,
                                 exit_p, pnl, pct, status, hit, dur)
    return {
        "status":      status,
        "target_hit":  hit,
        "closed_at":   _ist(trade.closed_at),
        "duration":    dur,
        "pnl_rs":      round(pnl, 0),
        "pnl_pct":     round(pct, 2),
        "expert_note": note,
    }

# ── Agent Trade row builders ──────────────────────────────────────────────────

def _agent_why_bought(trade: AgentTrade, decision: "AgentDecision | None") -> str:
    """Build the 'Why Bought' narrative for an agent trade."""
    parts = []
    parts.append(f"[AI Agent — {trade.strategy}]")
    parts.append(f"Regime: {trade.regime}")
    parts.append(f"Product: {trade.product}")
    if decision:
        parts.append(f"Confidence: {decision.confidence}%")
        if decision.reasons:
            for r in decision.reasons[:8]:   # cap to keep readable
                parts.append(f"• {r}")
    return "\n".join(parts)


def _agent_exit_status(exit_reason: str | None) -> tuple[str, str]:
    """Map agent exit_reason → (status label, target_hit label)."""
    r = (exit_reason or "").upper()
    if "SL_HIT" in r or "STOP_HIT" in r:
        return "STOPPED ❌", "Stop-loss triggered"
    if "T2_TARGET" in r:
        return "T2 HIT ✅", "T2 full target reached"
    if "T1_PARTIAL" in r:
        return "T1 HIT ✅", "T1 partial target reached"
    if "HUB_REVERSAL" in r:
        return "HUB EXIT ⚠️", f"Intelligence reversal: {exit_reason}"
    if "HUB_DETERIORATION" in r:
        return "HUB EXIT ⚠️", f"Score deteriorated: {exit_reason}"
    if "MIS_SQUAREOFF" in r:
        return "MIS EOD ✅", "Intraday square-off at 3:15 PM"
    if "MAX_HOLD" in r:
        return "MAX HOLD ⏰", "10-day max hold exceeded"
    if "DUPLICATE" in r:
        return "CANCELLED", "Duplicate cleanup"
    return "CLOSED", exit_reason or "Manual close"


def _agent_open_row(trade: AgentTrade, decision: "AgentDecision | None", hub) -> dict:
    entry   = trade.entry_price
    stop    = trade.stop_price
    t1      = trade.target_price
    risk    = abs(entry - stop)
    t2      = round(entry + 2 * (t1 - entry), 2) if trade.side == "BUY" else round(entry - 2 * (entry - t1), 2)
    atr     = round(risk / 2.0, 2)          # stop = entry ± 2×ATR → ATR = risk/2
    size_rs = round(trade.qty * entry, 0)
    max_risk_rs = round(trade.qty * risk, 0)
    conf    = decision.confidence if decision else 0
    hub_scores = _hub_scores(hub)
    hub_note = {k.replace("hub_", ""): v for k, v in hub_scores.items()} if hub_scores else None
    why = _agent_why_bought(trade, decision)
    note = build_expert_note(
        trade.symbol, trade.side, entry, stop, t1, t2, conf, hub_note, why
    )
    row = {
        "trade_id":    trade.id,          # UUID string — handled by updated existing_ids()
        "opened_at":   _ist(trade.entry_ts),
        "symbol":      trade.symbol.replace(".NS", ""),
        "direction":   trade.side,
        "qty":         trade.qty,
        "entry":       round(entry, 2),
        "stop":        round(stop, 2),
        "stop_pct":    f"{_pct(stop, entry):+.1f}%",
        "target_1":    round(t1, 2),
        "t1_pct":      f"{_pct(t1, entry):+.1f}%",
        "target_2":    round(t2, 2),
        "t2_pct":      f"{_pct(t2, entry):+.1f}%",
        "rr":          f"1:{_rr(entry, stop, t1)}",
        "atr":         atr,
        "confidence":  conf,
        "size_rs":     size_rs,
        "max_risk_rs": max_risk_rs,
        "why_bought":  note,
        "eta_t1":      estimate_eta_to_target(entry, t1, atr, trade.side),
        "status":      "OPEN",
    }
    row.update(hub_scores)
    row.update(_live_formulas())
    return row


def _agent_close_row(trade: AgentTrade) -> dict:
    exit_p = trade.exit_price or 0.0
    status, hit = _agent_exit_status(trade.exit_reason)
    dur  = _fmt_duration(trade.entry_ts, trade.exit_ts) if trade.exit_ts else ""
    pnl  = trade.pnl or 0.0
    pct  = trade.pnl_pct or 0.0
    note = build_postmortem_note(
        trade.symbol, trade.side, trade.entry_price,
        exit_p, pnl, pct, status, hit, dur,
    )
    return {
        "status":      status,
        "target_hit":  hit,
        "closed_at":   _ist(trade.exit_ts),
        "duration":    dur,
        "pnl_rs":      round(pnl, 0),
        "pnl_pct":     round(pct, 2),
        "expert_note": note,
    }


def _agent_hold_update(trade: AgentTrade, current_price: float,
                       pnl: float, pnl_pct: float, days_held: int,
                       hub) -> dict:
    """Refresh 'Expert Note' with live hold analysis for an open agent position."""
    hub_scores_raw = _hub_scores(hub) if hub else {}
    hub_note = {k.replace("hub_", ""): v for k, v in hub_scores_raw.items()} or None
    entry = trade.entry_price
    t1    = trade.target_price
    t2    = round(entry + 2 * (t1 - entry), 2) if trade.side == "BUY" \
            else round(entry - 2 * (entry - t1), 2)
    analysis = build_hold_analysis(
        symbol=trade.symbol,
        side=trade.side,
        entry=entry,
        current=current_price,
        stop=trade.stop_price,
        target_1=t1,
        target_2=t2,
        pnl=pnl,
        pnl_pct=pnl_pct,
        days_held=days_held,
        hub=hub_note,
        strategy=trade.strategy or "HUB_SIGNAL",
    )
    return {
        "expert_note": analysis,
    }


# ── Formatting helpers ────────────────────────────────────────────────────────
def _cell_fmt(bg=None, fg=None, bold=False, halign=None, wrap=False):
    fmt: dict = {}
    if bg:
        fmt["backgroundColor"] = bg
    tf: dict = {}
    if fg:
        tf["foregroundColor"] = fg
    if bold:
        tf["bold"] = True
    if tf:
        fmt["textFormat"] = tf
    if halign:
        fmt["horizontalAlignment"] = halign
    if wrap:
        fmt["wrapStrategy"] = "WRAP"
    return fmt

def _repeat(ws_id, r0, r1, c0, c1, fmt, fields):
    return {"repeatCell": {
        "range":  {"sheetId": ws_id, "startRowIndex": r0, "endRowIndex": r1,
                   "startColumnIndex": c0, "endColumnIndex": c1},
        "cell":   {"userEnteredFormat": fmt},
        "fields": f"userEnteredFormat({fields})",
    }}

def _cond_contains(ws_id, c0, c1, val, bg):
    return {"addConditionalFormatRule": {"rule": {
        "ranges": [{"sheetId": ws_id, "startRowIndex": 1,
                    "startColumnIndex": c0, "endColumnIndex": c1}],
        "booleanRule": {
            "condition": {"type": "TEXT_CONTAINS", "values": [{"userEnteredValue": val}]},
            "format": {"backgroundColor": bg},
        }
    }, "index": 0}}

def _cond_number(ws_id, c0, c1, op, val, bg):
    return {"addConditionalFormatRule": {"rule": {
        "ranges": [{"sheetId": ws_id, "startRowIndex": 1,
                    "startColumnIndex": c0, "endColumnIndex": c1}],
        "booleanRule": {
            "condition": {"type": op, "values": [{"userEnteredValue": str(val)}]},
            "format": {"backgroundColor": bg},
        }
    }, "index": 0}}


def _setup_trades_sheet(sh, ws):
    ws_id  = ws.id
    n_cols = len(KEYS)
    reqs   = []

    # Expand column count if needed
    reqs.append({"updateSheetProperties": {
        "properties": {"sheetId": ws_id,
                       "gridProperties": {"columnCount": n_cols + 4}},
        "fields": "gridProperties.columnCount",
    }})

    # Freeze header
    reqs.append({"updateSheetProperties": {
        "properties": {"sheetId": ws_id,
                       "gridProperties": {"frozenRowCount": 1}},
        "fields": "gridProperties.frozenRowCount",
    }})

    # Header row height
    reqs.append({"updateDimensionProperties": {
        "range": {"sheetId": ws_id, "dimension": "ROWS", "startIndex": 0, "endIndex": 1},
        "properties": {"pixelSize": 28}, "fields": "pixelSize",
    }})

    # Column widths
    for ci, px in enumerate(_COL_WIDTHS[:n_cols]):
        reqs.append({"updateDimensionProperties": {
            "range": {"sheetId": ws_id, "dimension": "COLUMNS",
                      "startIndex": ci, "endIndex": ci + 1},
            "properties": {"pixelSize": px}, "fields": "pixelSize",
        }})

    # Header row formatting (dark navy)
    reqs.append(_repeat(ws_id, 0, 1, 0, n_cols,
        _cell_fmt(bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, halign="CENTER"),
        "backgroundColor,textFormat,horizontalAlignment"))

    # Cyan highlight on the 4 live-data header cells
    live_start = _CI["live_price"]
    live_end   = _CI["days_held"] + 1
    reqs.append(_repeat(ws_id, 0, 1, live_start, live_end,
        _cell_fmt(bg=_rgb(6,182,212), fg=C_HEADER_BG, bold=True, halign="CENTER"),
        "backgroundColor,textFormat,horizontalAlignment"))

    # Center numeric/short columns in data rows
    center_keys = [
        "trade_id","direction","stop_pct","t1_pct","t2_pct","rr","atr",
        "confidence","size_rs","max_risk_rs",
        "hub_tech","hub_news","hub_sector","hub_macro","hub_earn","hub_fund","hub_opts",
        "live_price","live_pnl_rs","live_pnl_pct","days_held",
        "eta_t1","status","duration","pnl_rs","pnl_pct",
    ]
    for k in center_keys:
        ci = _CI[k]
        reqs.append(_repeat(ws_id, 1, 2000, ci, ci+1,
            _cell_fmt(halign="CENTER"), "horizontalAlignment"))

    # Wrap AI text columns
    for k in ("why_bought", "expert_note"):
        ci = _CI[k]
        reqs.append(_repeat(ws_id, 1, 2000, ci, ci+1,
            _cell_fmt(wrap=True), "wrapStrategy"))

    # Conditional: whole row color by direction
    reqs.append(_cond_contains(ws_id, 0, n_cols, "BUY",  C_BUY_BG))
    reqs.append(_cond_contains(ws_id, 0, n_cols, "SELL", C_SELL_BG))

    # Conditional: Status column
    ci_st = _CI["status"]
    reqs.append(_cond_contains(ws_id, ci_st, ci_st+1, "OPEN",    C_STATUS_OPEN))
    reqs.append(_cond_contains(ws_id, ci_st, ci_st+1, "T1 HIT",  C_STATUS_T1))
    reqs.append(_cond_contains(ws_id, ci_st, ci_st+1, "STOPPED", C_STATUS_STOP))

    # Conditional: Live P&L and Final P&L columns (green/red on positive/negative)
    for k in ("live_pnl_rs", "pnl_rs"):
        ci = _CI[k]
        reqs.append(_cond_number(ws_id, ci, ci+1, "NUMBER_GREATER", 0, C_PNL_POS))
        reqs.append(_cond_number(ws_id, ci, ci+1, "NUMBER_LESS",    0, C_PNL_NEG))
    for k in ("live_pnl_pct", "pnl_pct"):
        ci = _CI[k]
        reqs.append(_cond_number(ws_id, ci, ci+1, "NUMBER_GREATER", 0, C_PNL_POS))
        reqs.append(_cond_number(ws_id, ci, ci+1, "NUMBER_LESS",    0, C_PNL_NEG))

    # Conditional: Hub factor cells — positive=green, negative=red
    for k in ("hub_tech","hub_news","hub_sector","hub_macro","hub_earn","hub_fund","hub_opts"):
        ci = _CI[k]
        reqs.append(_cond_number(ws_id, ci, ci+1, "NUMBER_GREATER", 0, C_HUB_POS))
        reqs.append(_cond_number(ws_id, ci, ci+1, "NUMBER_LESS",    0, C_HUB_NEG))

    # Conditional: Days Held — highlight if held > 15 days (overdue)
    ci_days = _CI["days_held"]
    reqs.append(_cond_number(ws_id, ci_days, ci_days+1, "NUMBER_GREATER", 15,
                             _rgb(254, 243, 199)))   # amber warning

    # Auto-filter (filter dropdowns on every header column — lets user filter
    # by Status=OPEN/CLOSED, Direction=BUY/SELL, Symbol, etc.)
    reqs.append({"setBasicFilter": {
        "filter": {
            "range": {
                "sheetId": ws_id,
                "startRowIndex": 0,
                "startColumnIndex": 0,
                "endColumnIndex": n_cols,
            }
        }
    }})

    sh.batch_update({"requests": reqs})


def _write_daily_report(sh, daily_rows: list[dict]):
    """Create/refresh the '📅 Daily Report' sheet — one row per trading day.

    Columns: Date, Trades Opened, Trades Closed, Wins, Losses, Win %,
    Realised P&L, Best Trade, Worst Trade, Cumulative P&L.
    """
    from gspread.exceptions import WorksheetNotFound
    title = "📅 Daily Report"
    try:
        ws = sh.worksheet(title)
        ws.clear()
    except WorksheetNotFound:
        ws = sh.add_worksheet(title=title, rows=400, cols=10)

    headers = ["Date", "Opened", "Closed", "Wins", "Losses", "Win %",
               "Realised P&L", "Best Trade", "Worst Trade", "Cumulative P&L"]
    matrix = [headers]
    cum = 0.0
    for d in daily_rows:
        cum += d["realised"]
        matrix.append([
            d["date"], d["opened"], d["closed"], d["wins"], d["losses"],
            f'{d["win_rate"]:.0f}%',
            round(d["realised"], 2), round(d["best"], 2), round(d["worst"], 2),
            round(cum, 2),
        ])
    ws.update(matrix, "A1")
    try:
        ws.format("A1:J1", {"textFormat": {"bold": True},
                            "backgroundColor": {"red": 0.15, "green": 0.18, "blue": 0.28}})
    except Exception:
        pass


def _add_summary_sheet(sh, trades_ws_title="Trades"):
    """Create/refresh the 📊 Summary dashboard sheet."""
    try:
        summary = sh.worksheet("📊 Summary")
    except Exception:
        summary = sh.add_worksheet(title="📊 Summary", rows=60, cols=8)

    T   = trades_ws_title
    C   = _COL

    # Column letter shorthands for summary formulas
    dir_c  = C["direction"]      # D
    st_c   = C["status"]         # AD
    pnl_c  = C["pnl_rs"]        # AH
    pct_c  = C["pnl_pct"]       # AI
    conf_c = C["confidence"]     # N
    sym_c  = C["symbol"]         # C
    lpnl_c = C["live_pnl_rs"]   # AA
    lp_c   = C["live_price"]     # Z
    days_c = C["days_held"]      # AC

    rows = [
        # Row 1: Title
        ["🤖 AutoTrade Pro — Live Trade Journal", "", "", "", "", "", "", ""],
        [f"Refreshed:", f'=TEXT(NOW(),"dd mmm yyyy HH:MM")', "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", ""],

        # Row 4: Section header
        ["📊 PORTFOLIO SNAPSHOT", "", "", "", "📈 P&L SUMMARY", "", "", ""],
        ["Total Trades",
         f"=COUNTA('{T}'!A2:A)",
         "",
         "",
         "Total Realized P&L ₹",
         f'=IFERROR(SUMIF(\'{T}\'!{st_c}2:{st_c},"<>OPEN",\'{T}\'!{pnl_c}2:{pnl_c}),0)',
         "", ""],
        ["Open Positions",
         f"=COUNTIF('{T}'!{st_c}2:{st_c},\"OPEN\")",
         "",
         "",
         "Total Unrealized P&L ₹",
         f"=IFERROR(SUMIF('{T}'!{st_c}2:{st_c},\"OPEN\",'{T}'!{lpnl_c}2:{lpnl_c}),0)",
         "", ""],
        ["Total Closed",
         f"=COUNTIFS('{T}'!{st_c}2:{st_c},\"<>OPEN\",'{T}'!{st_c}2:{st_c},\"<>\")",
         "",
         "",
         "Best Trade ₹",
         f'=IFERROR(IF(COUNTIFS(\'{T}\'!{st_c}2:{st_c},"<>OPEN",\'{T}\'!{st_c}2:{st_c},"<>")=0,"—",MAX(\'{T}\'!{pnl_c}2:{pnl_c})),"—")',
         "", ""],
        ["Closed (Profit)",
         f"=COUNTIFS('{T}'!{st_c}2:{st_c},\"<>OPEN\",'{T}'!{pnl_c}2:{pnl_c},\">0\")",
         "",
         "",
         "Worst Trade ₹",
         f'=IFERROR(IF(COUNTIFS(\'{T}\'!{st_c}2:{st_c},"<>OPEN",\'{T}\'!{st_c}2:{st_c},"<>")=0,"—",MIN(\'{T}\'!{pnl_c}2:{pnl_c})),"—")',
         "", ""],
        ["Closed (Loss)",
         f"=COUNTIFS('{T}'!{st_c}2:{st_c},\"<>OPEN\",'{T}'!{pnl_c}2:{pnl_c},\"<0\")",
         "",
         "",
         "Avg P&L per trade ₹",
         f"=IFERROR(AVERAGEIF('{T}'!{st_c}2:{st_c},\"<>OPEN\",'{T}'!{pnl_c}2:{pnl_c}),\"—\")",
         "", ""],
        ["Win Rate",
         f"=IFERROR(TEXT(B8/(B8+B9),\"0.0%\"),\"—\")",
         "",
         "",
         "Avg Hold Duration (days)",
         f"=IFERROR(ROUND(AVERAGEIF('{T}'!{days_c}2:{days_c},\"<>—\",'{T}'!{days_c}2:{days_c}),1),\"—\")",
         "", ""],
        ["BUY trades",
         f"=COUNTIF('{T}'!{dir_c}2:{dir_c},\"BUY\")",
         "",
         "",
         "Avg Confidence %",
         f"=IFERROR(ROUND(AVERAGE('{T}'!{conf_c}2:{conf_c}),1),\"—\")",
         "", ""],
        ["SELL trades",
         f"=COUNTIF('{T}'!{dir_c}2:{dir_c},\"SELL\")",
         "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", ""],

        # Row 13: Live open positions
        ["⚡ OPEN POSITIONS (live prices auto-refresh)", "", "", "", "", "", "", ""],
        ["Symbol", "Dir", "Entry ₹", "Live ₹", "Live P&L ₹", "Live P&L %", "Days Held", "ETA"],
    ]

    # Last 8 open trades (most recent first) using FILTER+SORT
    eta_c = C["eta_t1"]
    entry_c = C["entry"]
    for i in range(1, 9):
        rows.append([
            f"=IFERROR(INDEX(FILTER('{T}'!{sym_c}$2:{sym_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"),COUNTA(FILTER('{T}'!{sym_c}$2:{sym_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"))+1-{i}),\"\")",
            f"=IFERROR(INDEX(FILTER('{T}'!{dir_c}$2:{dir_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"),COUNTA(FILTER('{T}'!{dir_c}$2:{dir_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"))+1-{i}),\"\")",
            f"=IFERROR(INDEX(FILTER('{T}'!{entry_c}$2:{entry_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"),COUNTA(FILTER('{T}'!{entry_c}$2:{entry_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"))+1-{i}),\"\")",
            f"=IFERROR(INDEX(FILTER('{T}'!{lp_c}$2:{lp_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"),COUNTA(FILTER('{T}'!{lp_c}$2:{lp_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"))+1-{i}),\"\")",
            f"=IFERROR(INDEX(FILTER('{T}'!{lpnl_c}$2:{lpnl_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"),COUNTA(FILTER('{T}'!{lpnl_c}$2:{lpnl_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"))+1-{i}),\"\")",
            f"=IFERROR(INDEX(FILTER('{T}'!{C['live_pnl_pct']}$2:{C['live_pnl_pct']},'{T}'!{st_c}$2:{st_c}=\"OPEN\"),COUNTA(FILTER('{T}'!{C['live_pnl_pct']}$2:{C['live_pnl_pct']},'{T}'!{st_c}$2:{st_c}=\"OPEN\"))+1-{i}),\"\")",
            f"=IFERROR(INDEX(FILTER('{T}'!{days_c}$2:{days_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"),COUNTA(FILTER('{T}'!{days_c}$2:{days_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"))+1-{i}),\"\")",
            f"=IFERROR(INDEX(FILTER('{T}'!{eta_c}$2:{eta_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"),COUNTA(FILTER('{T}'!{eta_c}$2:{eta_c},'{T}'!{st_c}$2:{st_c}=\"OPEN\"))+1-{i}),\"\")",
        ])

    rows.append(["", "", "", "", "", "", "", ""])

    # P&L chart section (SPARKLINE)
    rows.append(["📉 P&L CHART (closed trades)", "", "", "", "", "", "", ""])
    rows.append([
        f'=IFERROR(SPARKLINE(FILTER(\'{T}\'!{pnl_c}2:{pnl_c},\'{T}\'!{pnl_c}2:{pnl_c}<>""),'
        f'{{"charttype","column";"negcolor","#F43F5E";"color","#10B981";"ymin",'
        f'MIN(\'{T}\'!{pnl_c}2:{pnl_c});"ymax",MAX(\'{T}\'!{pnl_c}2:{pnl_c})}})'
        f',"No closed trades yet")',
        "", "", "", "", "", "", "",
    ])
    rows.append(["← Each bar = one closed trade | Green = profit, Red = loss", "", "", "", "", "", "", ""])

    summary.clear()
    summary.update(rows, "A1", value_input_option="USER_ENTERED")

    # Format the summary sheet
    ws_id = summary.id
    n = 8  # columns in summary
    sh.batch_update({"requests": [
        # Expand columns
        {"updateSheetProperties": {
            "properties": {"sheetId": ws_id, "gridProperties": {"columnCount": 10}},
            "fields": "gridProperties.columnCount"}},
        # Title row
        _repeat(ws_id, 0, 1, 0, n,
            _cell_fmt(bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True),
            "backgroundColor,textFormat"),
        # Section header rows bold
        _repeat(ws_id, 3, 4, 0, n,
            _cell_fmt(bg=_rgb(30,41,59), fg=C_HEADER_FG, bold=True),
            "backgroundColor,textFormat"),
        _repeat(ws_id, 13, 14, 0, n,
            _cell_fmt(bg=_rgb(6,78,59), fg=_rgb(209,250,229), bold=True),
            "backgroundColor,textFormat"),
        _repeat(ws_id, 25, 26, 0, n,
            _cell_fmt(bg=_rgb(30,41,59), fg=C_HEADER_FG, bold=True),
            "backgroundColor,textFormat"),
        # Column header row for open positions
        _repeat(ws_id, 14, 15, 0, n,
            _cell_fmt(bg=_rgb(51,65,85), fg=C_HEADER_FG, bold=True, halign="CENTER"),
            "backgroundColor,textFormat,horizontalAlignment"),
        # Open positions rows alternate shading
        _repeat(ws_id, 15, 23, 0, n,
            _cell_fmt(halign="CENTER"), "horizontalAlignment"),
        # Number formats — KPI value column (F, col 5): ₹ currency with commas (rows 5-12)
        {"repeatCell": {
            "range": {"sheetId": ws_id, "startRowIndex": 4, "endRowIndex": 12,
                      "startColumnIndex": 5, "endColumnIndex": 6},
            "cell": {"userEnteredFormat": {"numberFormat": {"type": "NUMBER", "pattern": "#,##0"}}},
            "fields": "userEnteredFormat.numberFormat"}},
        # Open positions: Live P&L ₹ (col E = idx 4) → comma-separated integer
        {"repeatCell": {
            "range": {"sheetId": ws_id, "startRowIndex": 15, "endRowIndex": 23,
                      "startColumnIndex": 4, "endColumnIndex": 5},
            "cell": {"userEnteredFormat": {"numberFormat": {"type": "NUMBER", "pattern": "#,##0"}}},
            "fields": "userEnteredFormat.numberFormat"}},
        # Open positions: Live P&L % (col F = idx 5) → 0.46 shown as "0.46%" (literal %)
        {"repeatCell": {
            "range": {"sheetId": ws_id, "startRowIndex": 15, "endRowIndex": 23,
                      "startColumnIndex": 5, "endColumnIndex": 6},
            "cell": {"userEnteredFormat": {"numberFormat": {"type": "NUMBER", "pattern": '0.00"%"'}}},
            "fields": "userEnteredFormat.numberFormat"}},
        # Open positions: Entry ₹ and Live ₹ (cols C,D = idx 2,3) → 2 decimal places
        {"repeatCell": {
            "range": {"sheetId": ws_id, "startRowIndex": 15, "endRowIndex": 23,
                      "startColumnIndex": 2, "endColumnIndex": 4},
            "cell": {"userEnteredFormat": {"numberFormat": {"type": "NUMBER", "pattern": "#,##0.00"}}},
            "fields": "userEnteredFormat.numberFormat"}},
        # Sparkline row height
        {"updateDimensionProperties": {
            "range": {"sheetId": ws_id, "dimension": "ROWS",
                      "startIndex": 26, "endIndex": 27},
            "properties": {"pixelSize": 140}, "fields": "pixelSize"}},
        # Column widths for summary
        {"updateDimensionProperties": {
            "range": {"sheetId": ws_id, "dimension": "COLUMNS",
                      "startIndex": 0, "endIndex": 1},
            "properties": {"pixelSize": 200}, "fields": "pixelSize"}},
        {"updateDimensionProperties": {
            "range": {"sheetId": ws_id, "dimension": "COLUMNS",
                      "startIndex": 1, "endIndex": 8},
            "properties": {"pixelSize": 130}, "fields": "pixelSize"}},
    ]})
    return summary


# ── Local Excel sink ──────────────────────────────────────────────────────────
class LocalExcelSink:
    def __init__(self, path: str):
        self.path = path

    def _load(self):
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        if os.path.exists(self.path):
            wb = openpyxl.load_workbook(self.path)
            ws = wb.active
        else:
            os.makedirs(os.path.dirname(self.path) or ".", exist_ok=True)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trades"
            ws.append(HEADERS)
            ws.freeze_panes = "A2"
            hdr_fill  = PatternFill("solid", fgColor="0F172A")
            hdr_font  = Font(bold=True, color="F8FAFC", size=10)
            hdr_align = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, hdr_align
            ws.row_dimensions[1].height = 22
            excel_widths = [
                6, 17, 13, 6, 10, 10, 7, 10, 7, 10, 7, 6, 7,
                7, 11, 12, 6, 6, 7, 7, 6, 6, 6,
                50, 17, 10, 13, 9, 9,
                10, 22, 17, 10, 11, 8, 50,
            ]
            for ci, w in enumerate(excel_widths[:len(KEYS)], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
        return wb, ws

    def existing_ids(self) -> dict:
        if not os.path.exists(self.path):
            return {}
        import openpyxl
        wb = openpyxl.load_workbook(self.path, read_only=True)
        ws = wb.active
        ci_st = _CI["status"]
        out = {}
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[0] in (None, ""):
                continue
            # Accept both int IDs (paper trades) and string IDs (agent trades)
            tid = row[0] if isinstance(row[0], str) else None
            if tid is None:
                try:
                    tid = int(row[0])
                except (ValueError, TypeError):
                    continue
            status = row[ci_st] if len(row) > ci_st else ""
            out[tid] = (i, status or "")
        wb.close()
        return out

    def append(self, row: dict):
        import openpyxl
        from openpyxl.styles import PatternFill, Alignment
        direction = row.get("direction", "")
        with _FILE_LOCK:
            wb, ws = self._load()
            # Skip formula strings in Excel
            values = []
            for k in KEYS:
                v = row.get(k, "")
                if isinstance(v, str) and v.startswith("="):
                    v = ""
                values.append(v)
            ws.append(values)
            ri = ws.max_row
            if direction == "BUY":
                fill = PatternFill("solid", fgColor="F0FDF4")
            elif direction == "SELL":
                fill = PatternFill("solid", fgColor="FFF1F2")
            else:
                fill = None
            if fill:
                for cell in ws[ri]:
                    cell.fill = fill
            for k in ("why_bought", "expert_note"):
                ws.cell(ri, _CI[k]+1).alignment = Alignment(wrap_text=True, vertical="top")
            wb.save(self.path)

    def update(self, row_number: int, partial: dict):
        import openpyxl
        from openpyxl.styles import PatternFill
        with _FILE_LOCK:
            wb, ws = self._load()
            for k, v in partial.items():
                if k in KEYS:
                    ws.cell(row=row_number, column=_CI[k]+1, value=v)
            pnl = partial.get("pnl_rs")
            if pnl is not None:
                fill = PatternFill("solid", fgColor="D1FAE5" if float(pnl) >= 0 else "FEE2E2")
                ws.cell(row_number, _CI["pnl_rs"]+1).fill = fill
            status = partial.get("status", "")
            if status:
                fill = PatternFill("solid",
                    fgColor="D1FAE5" if "T1 HIT" in status else
                            "FEE2E2" if "STOPPED" in status else "F8FAFC")
                ws.cell(row_number, _CI["status"]+1).fill = fill
            wb.save(self.path)

    def write_summary_sheet(self, open_trades: list[dict], closed_trades: list[dict]):
        """Create/replace the '📊 Portfolio Summary' sheet with three sections.

        Each section has: overview stats, HOLDING (open), PROFIT, LOSS.
        open_trades and closed_trades are dicts with keys from SCHEMA plus
        extra 'current_price', 'pnl', 'pnl_pct', 'hold_note', 'expert_note'.
        """
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        with _FILE_LOCK:
            # Load or create the main workbook
            if os.path.exists(self.path):
                wb = openpyxl.load_workbook(self.path)
            else:
                wb = openpyxl.Workbook()

            # Remove old summary sheet if it exists
            if "📊 Portfolio Summary" in wb.sheetnames:
                del wb["📊 Portfolio Summary"]

            ws = wb.create_sheet("📊 Portfolio Summary", 0)   # first tab

            # ── Styles ───────────────────────────────────────────────────────────
            def hdr_style(bg, fg="FFFFFF", bold=True, sz=11):
                return Font(bold=bold, color=fg, size=sz), PatternFill("solid", fgColor=bg)

            title_font, title_fill   = hdr_style("0F172A", "94E3FF", sz=14)
            sec_font, sec_fill_hold  = hdr_style("064E3B", "A7F3D0", sz=11)
            sec_font2, sec_fill_prof = hdr_style("14532D", "BBF7D0", sz=11)
            sec_font3, sec_fill_loss = hdr_style("7F1D1D", "FECACA", sz=11)
            col_font, col_fill       = hdr_style("1E293B", "CBD5E1", sz=9)
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            wrap   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
            thin_border = Border(
                bottom=Side(style="thin", color="475569"),
            )

            profits = [t for t in closed_trades if (t.get("pnl") or 0) >= 0]
            losses  = [t for t in closed_trades if (t.get("pnl") or 0) < 0]
            total_pnl = sum((t.get("pnl") or 0) for t in closed_trades)
            unrealised = sum((t.get("pnl") or 0) for t in open_trades)
            start_capital = settings.AGENT_EQUITY
            equity  = start_capital + total_pnl + unrealised
            roi_pct = (equity - start_capital) / start_capital * 100
            win_rate = len(profits) / max(1, len(closed_trades)) * 100

            # ── Overview block ────────────────────────────────────────────────────
            row = 1
            ws.merge_cells(f"A{row}:H{row}")
            ws[f"A{row}"] = "📊 AUTOTRADE PRO — PORTFOLIO SUMMARY"
            ws[f"A{row}"].font  = title_font
            ws[f"A{row}"].fill  = title_fill
            ws[f"A{row}"].alignment = center
            ws.row_dimensions[row].height = 26
            row += 1

            from datetime import datetime
            ws.merge_cells(f"A{row}:H{row}")
            ws[f"A{row}"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}   |   Agent Solo Mode   |   ₹{start_capital/100000:.0f}L Capital"
            ws[f"A{row}"].font = Font(color="64748B", italic=True, size=9)
            ws[f"A{row}"].alignment = center
            row += 2

            kpis = [
                ("Portfolio Equity",   f"₹{equity:,.0f}",           "94FAD0"),
                ("Start Capital",      f"₹{start_capital:,.0f}",    "E2E8F0"),
                ("Total Realised P&L", f"₹{total_pnl:+,.0f}",       "BBF7D0" if total_pnl >= 0 else "FECACA"),
                ("Unrealised P&L",     f"₹{unrealised:+,.0f}",      "BBF7D0" if unrealised >= 0 else "FECACA"),
                ("ROI",                f"{roi_pct:+.2f}%",           "BBF7D0" if roi_pct >= 0 else "FECACA"),
                ("Win Rate",           f"{win_rate:.1f}%",           "BBF7D0" if win_rate >= 50 else "FEF3C7"),
                ("Open Positions",     str(len(open_trades)),        "DBEAFE"),
                ("Closed Trades",      str(len(closed_trades)),      "E0E7FF"),
            ]
            for col_idx, (label, value, color) in enumerate(kpis, 1):
                cl = get_column_letter(col_idx)
                ws[f"{cl}{row}"]   = label
                ws[f"{cl}{row}"].font = Font(bold=True, size=8, color="64748B")
                ws[f"{cl}{row}"].alignment = center
                ws.row_dimensions[row].height = 14
                ws[f"{cl}{row+1}"] = value
                ws[f"{cl}{row+1}"].font = Font(bold=True, size=12, color="0F172A")
                ws[f"{cl}{row+1}"].fill = PatternFill("solid", fgColor=color)
                ws[f"{cl}{row+1}"].alignment = center
                ws.row_dimensions[row+1].height = 22

            row += 3   # skip label + value + blank

            # ── Section builder helper ──────────────────────────────────────────
            def write_section(title, trades, header_fill_pair, cols, row_color):
                nonlocal row
                ws.merge_cells(f"A{row}:H{row}")
                ws[f"A{row}"] = title
                ws[f"A{row}"].font = header_fill_pair[0]
                ws[f"A{row}"].fill = header_fill_pair[1]
                ws[f"A{row}"].alignment = center
                ws.row_dimensions[row].height = 20
                row += 1

                # column headers
                for ci, (col_key, col_label) in enumerate(cols, 1):
                    cl = get_column_letter(ci)
                    ws[f"{cl}{row}"] = col_label
                    ws[f"{cl}{row}"].font = col_font
                    ws[f"{cl}{row}"].fill = col_fill
                    ws[f"{cl}{row}"].alignment = center
                    ws.row_dimensions[row].height = 14
                row += 1

                if not trades:
                    ws.merge_cells(f"A{row}:H{row}")
                    ws[f"A{row}"] = "No trades in this category."
                    ws[f"A{row}"].font = Font(italic=True, color="94A3B8", size=9)
                    row += 2
                    return

                for t in trades:
                    for ci, (col_key, _) in enumerate(cols, 1):
                        cl = get_column_letter(ci)
                        val = t.get(col_key, "")
                        ws[f"{cl}{row}"] = val
                        cell = ws[f"{cl}{row}"]
                        cell.fill = PatternFill("solid", fgColor=row_color)
                        if col_key in ("hold_note", "expert_note", "why_bought"):
                            cell.alignment = wrap
                            cell.font = Font(size=8)
                        else:
                            cell.alignment = center
                            cell.font = Font(size=9)
                        cell.border = thin_border
                    ws.row_dimensions[row].height = max(80, min(200,
                        len(str(t.get("hold_note") or t.get("expert_note", "")).split('\n')) * 14))
                    row += 1
                row += 1   # blank gap between sections

            # ── HOLDING section ───────────────────────────────────────────────────
            hold_cols = [
                ("symbol",       "Symbol"),
                ("direction",    "Dir"),
                ("entry",        "Entry ₹"),
                ("current_price","Current ₹"),
                ("pnl",          "Live P&L ₹"),
                ("pnl_pct",      "P&L %"),
                ("days_held",    "Days"),
                ("hold_note",    "⏳ Why Still Holding — Expert Analysis"),
            ]
            write_section(
                f"⏳  CURRENTLY HOLDING  ({len(open_trades)} positions)",
                sorted(open_trades, key=lambda t: (t.get("pnl") or 0), reverse=True),
                (sec_font, sec_fill_hold),
                hold_cols, "F0FDF4",
            )

            # ── PROFIT section ───────────────────────────────────────────────────
            prof_cols = [
                ("symbol",       "Symbol"),
                ("direction",    "Dir"),
                ("entry",        "Entry ₹"),
                ("exit_price",   "Exit ₹"),
                ("pnl",          "P&L ₹"),
                ("pnl_pct",      "P&L %"),
                ("duration",     "Held"),
                ("expert_note",  "✅ Why It Made Profit — Expert Analysis"),
            ]
            write_section(
                f"✅  CLOSED WITH PROFIT  ({len(profits)} trades  |  Total: ₹{sum(t.get('pnl',0) for t in profits):+,.0f})",
                sorted(profits, key=lambda t: (t.get("pnl") or 0), reverse=True),
                (sec_font2, sec_fill_prof),
                prof_cols, "DCFCE7",
            )

            # ── LOSS section ─────────────────────────────────────────────────────
            loss_cols = [
                ("symbol",       "Symbol"),
                ("direction",    "Dir"),
                ("entry",        "Entry ₹"),
                ("exit_price",   "Exit ₹"),
                ("pnl",          "P&L ₹"),
                ("pnl_pct",      "P&L %"),
                ("duration",     "Held"),
                ("expert_note",  "❌ Why It Took a Loss — Expert Analysis"),
            ]
            write_section(
                f"❌  CLOSED WITH LOSS  ({len(losses)} trades  |  Total: ₹{sum(t.get('pnl',0) for t in losses):+,.0f})",
                sorted(losses, key=lambda t: (t.get("pnl") or 0)),
                (sec_font3, sec_fill_loss),
                loss_cols, "FEF2F2",
            )

            # Column widths for summary sheet
            for ci, w in enumerate([14, 7, 12, 12, 14, 10, 8, 70], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

            wb.save(self.path)


# ── Google Sheets sink ────────────────────────────────────────────────────────
class GoogleSheetsSink:
    def __init__(self):
        self._ws = self._sh = None
        self._formatted = False
        self._row_count: int | None = None   # cached after existing_ids()
        self._pending_rows: list[list] = []  # batched appends — flushed once in flush()
        self._pending_updates: list[dict] = []  # batched cell updates for batch_update()

    def _get_credentials(self):
        import pickle
        from google.auth.transport.requests import Request
        token_path = settings.GOOGLE_OAUTH_TOKEN_PATH
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = None
        if os.path.exists(token_path):
            with open(token_path, "rb") as f:
                creds = pickle.load(f)
        if creds and creds.valid:
            return creds
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
                with open(token_path, "wb") as f:
                    pickle.dump(creds, f)
                return creds
            except Exception:
                pass
        secret = settings.GOOGLE_OAUTH_CLIENT_SECRET_JSON
        if secret and os.path.exists(secret):
            from google_auth_oauthlib.flow import InstalledAppFlow
            flow = InstalledAppFlow.from_client_secrets_file(secret, scopes)
            creds = flow.run_local_server(port=0)
            os.makedirs(os.path.dirname(token_path) or ".", exist_ok=True)
            with open(token_path, "wb") as f:
                pickle.dump(creds, f)
            return creds
        if settings.GOOGLE_SERVICE_ACCOUNT_JSON:
            from google.oauth2.service_account import Credentials
            return Credentials.from_service_account_file(
                settings.GOOGLE_SERVICE_ACCOUNT_JSON, scopes=scopes)
        raise RuntimeError("No Google credentials configured")

    def _worksheet(self):
        if self._ws is not None:
            return self._ws, self._sh
        import gspread
        from gspread.exceptions import WorksheetNotFound
        gc = gspread.authorize(self._get_credentials())
        sh = gc.open_by_key(settings.GOOGLE_SHEETS_ID)
        title = settings.GOOGLE_SHEETS_WORKSHEET

        # 1. Get-or-create the worksheet. Only create on WorksheetNotFound — a
        #    broad except here caused transient Google 503s during row_values()/
        #    update() to be mistaken for "sheet missing" → addSheet → "already
        #    exists" 400. Header maintenance is done in its own guarded block.
        try:
            ws = sh.worksheet(title)
        except WorksheetNotFound:
            ws = sh.add_worksheet(title=title, rows=2000, cols=len(KEYS) + 4)
            ws.update([HEADERS], "A1")
            self._formatted = False
            self._ws, self._sh = ws, sh
            return ws, sh

        # 2. Header maintenance (best-effort — never re-creates the sheet).
        try:
            existing = ws.row_values(1) if ws.row_count > 0 else []
            if not existing:
                ws.update([HEADERS], "A1")
                self._formatted = False
            elif existing != HEADERS:
                ws.clear()
                ws.update([HEADERS], "A1")
                self._formatted = False
        except Exception as exc:
            logger.debug(f"[sheet_logger] header check skipped (transient): {exc}")

        self._ws, self._sh = ws, sh
        return ws, sh

    def _ensure_formatted(self, ws, sh):
        if self._formatted:
            return
        try:
            _setup_trades_sheet(sh, ws)
            _add_summary_sheet(sh, ws.title)
            self._formatted = True
        except Exception as exc:
            logger.warning(f"[sheet_logger] formatting failed (non-fatal): {exc}")

    def existing_ids(self) -> dict:
        ws, sh = self._worksheet()
        self._ensure_formatted(ws, sh)
        records = ws.get_all_values()
        self._row_count = len(records)   # cache so append() doesn't re-fetch
        ci_st = _CI["status"]
        out = {}
        for i, row in enumerate(records[1:], start=2):
            if not row or not row[0]:
                continue
            raw = row[0]
            # Accept both int IDs (paper trades) and string IDs (agent trades)
            try:
                tid = int(raw)
            except (ValueError, TypeError):
                tid = str(raw)   # agent trade UUID
            status = row[ci_st] if len(row) > ci_st else ""
            out[tid] = (i, status or "")
        return out

    def append(self, row: dict):
        """Stage a row for batched writing. Call flush() after all appends."""
        if self._row_count is None:
            ws, _ = self._worksheet()
            self._row_count = len(ws.get_all_values())
        # Pre-allocate the row number so formulas are correct
        self._row_count += 1
        r = self._row_count
        values = []
        for k in KEYS:
            v = row.get(k, "")
            if isinstance(v, str) and "{ROW}" in v:
                v = v.replace("{ROW}", str(r))
            values.append(v)
        self._pending_rows.append(values)

    def flush(self):
        """Write all staged rows + cell updates in as few API calls as possible."""
        if self._pending_rows:
            ws, _ = self._worksheet()
            try:
                ws.append_rows(self._pending_rows, value_input_option="USER_ENTERED")
            except Exception as exc:
                logger.warning(f"[sheet_logger] batch append failed: {exc}")
            self._pending_rows.clear()

        if self._pending_updates:
            ws, sh = self._worksheet()
            try:
                sh.batch_update({"requests": self._pending_updates})
            except Exception as exc:
                # Fallback: update cells one by one
                for req in self._pending_updates:
                    try:
                        ws.batch_update(req)
                    except Exception:
                        pass
                logger.warning(f"[sheet_logger] batch cell update failed: {exc}")
            self._pending_updates.clear()

    def update(self, row_number: int, partial: dict):
        """Stage cell updates for batched writing. Flushed by flush()."""
        ws, _ = self._worksheet()
        for k, v in partial.items():
            if k not in _CI:
                continue
            # Do not overwrite live formulas in Google Sheets.
            if k in {"live_price", "live_pnl_rs", "live_pnl_pct", "days_held"}:
                continue
            col = _CI[k] + 1  # 1-based
            # Numeric fields must use numberValue so SUMIF/COUNTIFS treat them as
            # numbers — stringValue makes ">0" comparisons silently return 0.
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cell_val = {"numberValue": float(v)}
            else:
                cell_val = {"stringValue": str(v)}
            self._pending_updates.append({
                "updateCells": {
                    "range": {
                        "sheetId": ws.id,
                        "startRowIndex": row_number - 1,
                        "endRowIndex":   row_number,
                        "startColumnIndex": col - 1,
                        "endColumnIndex":   col,
                    },
                    "rows": [{"values": [{"userEnteredValue": cell_val}]}],
                    "fields": "userEnteredValue",
                }
            })


# ── Factory + sync engine ─────────────────────────────────────────────────────
def _make_sink():
    backend = (getattr(settings, "SHEET_LOG_BACKEND", "local") or "local").lower()
    if backend == "google":
        if not settings.google_sheets_available:
            logger.warning("[sheet_logger] google backend: no credentials — skipping")
            return None
        return GoogleSheetsSink()
    return LocalExcelSink(settings.SHEET_LOG_LOCAL_PATH)


def _sync_blocking(open_rows, close_updates):
    sink = _make_sink()
    if sink is None:
        return 0, 0
    appended = updated = 0
    for row in open_rows:
        sink.append(row)   # staged, not written yet
        appended += 1
    for row_number, partial in close_updates:
        sink.update(row_number, partial)   # staged
        updated += 1
    # Write all staged rows + updates in batch (1-2 API calls regardless of row count)
    if hasattr(sink, "flush"):
        try:
            sink.flush()
        except Exception as exc:
            logger.warning(f"[sheet_logger] batch flush failed: {exc}")
    return appended, updated


def _rebuild_summary(agent_rows, pos_map, hub_map):
    """Build the Portfolio Summary sheet from agent trade data.

    Runs in a thread — called via asyncio.to_thread from sync_journal.
    """
    sink = _make_sink()
    if not isinstance(sink, LocalExcelSink):
        return

    now = datetime.now(tz=timezone.utc)
    open_trades, closed_trades = [], []

    for t in agent_rows:
        if t.exit_price is None:
            pos = pos_map.get(t.symbol) or pos_map.get(t.symbol.replace(".NS", ""))
            current = (pos.current_price if pos else None) or t.entry_price
            raw_pnl = pos.unrealized_pnl if pos else 0.0
            pnl_pct = (raw_pnl / (t.entry_price * t.qty) * 100) if (t.entry_price and t.qty) else 0.0
            days    = max(0, (now - t.entry_ts.replace(tzinfo=timezone.utc if t.entry_ts.tzinfo is None else t.entry_ts.tzinfo)).days)
            hub     = hub_map.get(f"{t.symbol.split('.')[0]}.NS")
            hub_scores_raw = _hub_scores(hub) if hub else {}
            hub_note = {k.replace("hub_", ""): v for k, v in hub_scores_raw.items()} or None
            entry  = t.entry_price
            t1     = t.target_price
            t2     = round(entry + 2 * (t1 - entry), 2) if t.side == "BUY" else round(entry - 2 * (entry - t1), 2)
            hold_note = build_hold_analysis(
                symbol=t.symbol, side=t.side, entry=entry, current=current,
                stop=t.stop_price, target_1=t1, target_2=t2,
                pnl=raw_pnl, pnl_pct=pnl_pct, days_held=days,
                hub=hub_note, strategy=t.strategy or "HUB_SIGNAL",
            )
            open_trades.append({
                "symbol":        t.symbol.replace(".NS", ""),
                "direction":     t.side,
                "entry":         round(entry, 2),
                "current_price": round(current, 2),
                "pnl":           round(raw_pnl, 0),
                "pnl_pct":       f"{pnl_pct:+.2f}%",
                "days_held":     days,
                "hold_note":     hold_note,
            })
        else:
            dur = _fmt_duration(t.entry_ts, t.exit_ts) if t.exit_ts else ""
            pnl = t.pnl or 0.0
            pct = t.pnl_pct or 0.0
            status, hit = _agent_exit_status(t.exit_reason)
            expert = build_postmortem_note(
                t.symbol, t.side, t.entry_price, t.exit_price,
                pnl, pct, status, hit, dur,
            )
            closed_trades.append({
                "symbol":       t.symbol.replace(".NS", ""),
                "direction":    t.side,
                "entry":        round(t.entry_price, 2),
                "exit_price":   round(t.exit_price, 2),
                "pnl":          round(pnl, 0),
                "pnl_pct":      f"{pct:+.2f}%",
                "duration":     dur,
                "expert_note":  expert,
            })

    sink.write_summary_sheet(open_trades, closed_trades)


async def _compute_daily_report(session: AsyncSession) -> list[dict]:
    """Per-day aggregates from paper_trades: opened/closed/wins/losses/realised/best/worst."""
    from sqlalchemy import text as _text
    rows = (await session.execute(_text("""
        WITH opened AS (
            SELECT opened_at::date AS d, COUNT(*) AS n
            FROM paper_trades GROUP BY 1
        ),
        closed AS (
            SELECT closed_at::date AS d,
                   COUNT(*) AS n,
                   COUNT(*) FILTER (WHERE pnl > 0) AS wins,
                   COUNT(*) FILTER (WHERE pnl < 0) AS losses,
                   COALESCE(SUM(pnl), 0) AS realised,
                   COALESCE(MAX(pnl), 0) AS best,
                   COALESCE(MIN(pnl), 0) AS worst
            FROM paper_trades WHERE closed_at IS NOT NULL AND pnl IS NOT NULL
            GROUP BY 1
        )
        SELECT COALESCE(o.d, c.d) AS day,
               COALESCE(o.n, 0) AS opened,
               COALESCE(c.n, 0) AS closed,
               COALESCE(c.wins, 0) AS wins,
               COALESCE(c.losses, 0) AS losses,
               COALESCE(c.realised, 0) AS realised,
               COALESCE(c.best, 0) AS best,
               COALESCE(c.worst, 0) AS worst
        FROM opened o FULL OUTER JOIN closed c ON o.d = c.d
        ORDER BY day
    """))).all()
    out = []
    for r in rows:
        wl = (r.wins or 0) + (r.losses or 0)
        out.append({
            "date": str(r.day), "opened": r.opened, "closed": r.closed,
            "wins": r.wins, "losses": r.losses,
            "win_rate": (r.wins / wl * 100) if wl else 0.0,
            "realised": float(r.realised), "best": float(r.best), "worst": float(r.worst),
        })
    return out


async def sync_journal(session: AsyncSession, *, limit: int = 500) -> dict:
    """Idempotently reconcile the spreadsheet with both PaperTrade and AgentTrade tables."""
    if not getattr(settings, "SHEET_LOG_ENABLED", False):
        return {"enabled": False}
    try:
        sink = _make_sink()
        if sink is None:
            return {"enabled": True, "skipped": "no_sink"}

        existing = await asyncio.to_thread(sink.existing_ids)

        # ── Paper trades ──────────────────────────────────────────────────────
        paper_rows = (await session.execute(
            select(PaperTrade).order_by(PaperTrade.opened_at.desc()).limit(limit)
        )).scalars().all()

        # ── Agent trades ──────────────────────────────────────────────────────
        # The agent executor now writes every trade to paper_trades too (the
        # canonical table), so logging agent_trades here as well would DOUBLE
        # each trade in the sheet (PaperTrade.id is int, AgentTrade.id is uuid →
        # the dedup set never matches them). Log paper_trades only.
        agent_rows = []

        # Batch-fetch Hub scores for all symbols that need a new row
        need_hub_syms: set[str] = set()
        for t in paper_rows:
            if t.id not in existing:
                need_hub_syms.add(f"{t.symbol.split('.')[0]}.NS")
        for t in agent_rows:
            if t.id not in existing:
                need_hub_syms.add(f"{t.symbol.split('.')[0]}.NS")

        hub_map: dict = {}
        if need_hub_syms:
            for m in (await session.execute(
                select(MasterIntelligenceScore)
                .where(MasterIntelligenceScore.symbol.in_(need_hub_syms))
                .order_by(MasterIntelligenceScore.scored_at.desc())
            )).scalars().all():
                hub_map.setdefault(m.symbol, m)

        # Batch-fetch AgentDecisions for agent trades that need a new row
        new_agent_ids = [t.decision_id for t in agent_rows
                         if t.id not in existing and t.decision_id]
        decision_map: dict[str, AgentDecision] = {}
        if new_agent_ids:
            for d in (await session.execute(
                select(AgentDecision).where(AgentDecision.id.in_(new_agent_ids))
            )).scalars().all():
                decision_map[d.id] = d

        # ── Open agent positions — for live P&L and hold analysis ─────────────
        open_positions = (await session.execute(
            select(OpenPosition)
        )).scalars().all()

        pos_map = {p.symbol: p for p in open_positions}       # symbol → OpenPosition

        # Hub scores for hold-analysis refresh (open agent trades in existing sheet)
        all_open_syms = {f"{t.symbol.split('.')[0]}.NS"
                         for t in agent_rows if t.id in existing and t.exit_price is None}
        all_open_syms.update(need_hub_syms)
        if all_open_syms:
            for m in (await session.execute(
                select(MasterIntelligenceScore)
                .where(MasterIntelligenceScore.symbol.in_(all_open_syms))
                .order_by(MasterIntelligenceScore.scored_at.desc())
            )).scalars().all():
                hub_map.setdefault(m.symbol, m)

        open_rows, close_updates = [], []

        # Paper trades
        for t in paper_rows:
            if t.id not in existing:
                hub = hub_map.get(f"{t.symbol.split('.')[0]}.NS")
                row = _open_row(t, hub)
                # If the trade is ALREADY closed when first logged (e.g. after a
                # sheet rebuild), merge the close fields in so it isn't stuck OPEN.
                if t.status != TradeStatus.OPEN or t.exit_price is not None:
                    row.update(_close_partial(t))
                open_rows.append(row)
            else:
                row_number, sheet_status = existing[t.id]
                if "OPEN" in str(sheet_status) and t.status != TradeStatus.OPEN:
                    close_updates.append((row_number, _close_partial(t)))

        # Agent trades
        now = datetime.now(tz=timezone.utc)
        for t in agent_rows:
            if t.id not in existing:
                hub      = hub_map.get(f"{t.symbol.split('.')[0]}.NS")
                decision = decision_map.get(t.decision_id) if t.decision_id else None
                open_rows.append(_agent_open_row(t, decision, hub))
            else:
                row_number, sheet_status = existing[t.id]
                if "OPEN" in str(sheet_status):
                    if t.exit_price is not None:
                        close_updates.append((row_number, _agent_close_row(t)))
                    else:
                        # Refresh hold analysis for still-open positions
                        pos = pos_map.get(t.symbol) or pos_map.get(t.symbol.replace(".NS", ""))
                        current_price = (pos.current_price if pos else None) or t.entry_price
                        raw_pnl  = pos.unrealized_pnl if pos else 0.0
                        pnl_pct  = (raw_pnl / (t.entry_price * t.qty) * 100) if (t.entry_price and t.qty) else 0.0
                        days     = max(0, (now - t.entry_ts.replace(tzinfo=timezone.utc if t.entry_ts.tzinfo is None else t.entry_ts.tzinfo)).days)
                        hub      = hub_map.get(f"{t.symbol.split('.')[0]}.NS")
                        close_updates.append((row_number, _agent_hold_update(
                            t, current_price, raw_pnl, pnl_pct, days, hub
                        )))

        if not open_rows and not close_updates:
            # Still write the summary sheet even if no new/close changes
            if getattr(settings, "SHEET_LOG_BACKEND", "local") == "local":
                try:
                    await asyncio.to_thread(_rebuild_summary, agent_rows, pos_map, hub_map)
                except Exception as exc:
                    logger.warning(f"[sheet_logger] summary sheet failed: {exc}")
            return {"enabled": True, "appended": 0, "updated": 0,
                    "in_sheet": len(existing)}

        appended, updated = await asyncio.to_thread(_sync_blocking, open_rows, close_updates)
        logger.info(
            f"[sheet_logger] sync: +{appended} new "
            f"({len(paper_rows)} paper + {len(agent_rows)} agent), "
            f"~{updated} closed (backend={settings.SHEET_LOG_BACKEND})"
        )

        # Rebuild the portfolio summary sheet (local backend only)
        if getattr(settings, "SHEET_LOG_BACKEND", "local") == "local":
            try:
                await asyncio.to_thread(_rebuild_summary, agent_rows, pos_map, hub_map)
            except Exception as exc:
                logger.warning(f"[sheet_logger] summary sheet failed (non-fatal): {exc}")

        # ── Per-day report (both backends) ───────────────────────────────────
        try:
            daily_rows = await _compute_daily_report(session)
            if daily_rows and getattr(settings, "SHEET_LOG_BACKEND", "local") == "google":
                await asyncio.to_thread(_write_daily_report, sink._sh, daily_rows)
        except Exception as exc:
            logger.warning(f"[sheet_logger] daily report failed (non-fatal): {exc}")

        return {"enabled": True, "appended": appended, "updated": updated,
                "agent_trades": len(agent_rows), "paper_trades": len(paper_rows)}
    except Exception as exc:
        logger.warning(f"[sheet_logger] sync_journal failed (non-fatal): {exc}")
        return {"enabled": True, "error": str(exc)}
